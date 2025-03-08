import os
from pathlib import Path
from dotenv import load_dotenv
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
    ConversationHandler,
)
import re
from datetime import datetime, timedelta
from sqlalchemy import func
import csv
from io import StringIO, BytesIO
from logger import bot_logger
from database import SessionLocal
from models import User, Transaction, Category, TransactionType
from messages import *  # Импортируем все сообщения
import asyncio
from middleware import LoggingMiddleware, MetricsMiddleware
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Проверяем наличие .env файла
env_file = Path(".env")
if not env_file.exists():
    raise FileNotFoundError(SYSTEM_ENV_FILE_NOT_FOUND)

# Проверяем права доступа к .env файлу
if os.name != "nt":  # Пропускаем проверку для Windows
    env_permissions = oct(os.stat(env_file).st_mode)[-3:]
    if env_permissions != "600":
        raise PermissionError(
            SYSTEM_ENV_FILE_PERMISSIONS.format(permissions=env_permissions)
        )

# Загружаем переменные окружения
load_dotenv()

# Проверяем наличие необходимых переменных окружения
required_env_vars = ["TELEGRAM_BOT_TOKEN"]

missing_vars = [var for var in required_env_vars if not os.getenv(var)]
if missing_vars:
    raise EnvironmentError(SYSTEM_ENV_VARS_MISSING.format(vars=", ".join(missing_vars)))

# Настраиваем логгер
logger = bot_logger


class FinanceBot:
    def __init__(self):
        """Инициализация бота"""
        self.token = os.getenv("TELEGRAM_BOT_TOKEN")
        if not self.token:
            raise ValueError(SYSTEM_TOKEN_NOT_SET)

        self.admin_ids = [
            int(id) for id in os.getenv("ADMIN_USER_IDS", "").split(",") if id
        ]

        # Инициализируем middleware
        self.logging_middleware = LoggingMiddleware()
        self.metrics_middleware = MetricsMiddleware()

        # Регулярные выражения для парсинга сообщений
        self.expense_pattern = re.compile(
            r"^-\s*(\d+(?:[.,]\d+)?(?:\s*\d+)*)\s*(?:руб(?:лей|\.)?|р\.)?\s*(.+)$"
        )
        self.income_pattern = re.compile(
            r"^\+\s*(\d+(?:[.,]\d+)?(?:\s*\d+)*)\s*(?:руб(?:лей|\.)?|р\.)?\s*(.+)$"
        )

        # Регулярное выражение для поиска ключевых слов категорий
        self.category_keywords = {
            "Продукты": [
                "продукты",
                "еда",
                "магазин",
                "супермаркет",
                "вкусно и точка",
                "ростикс",
                "обед",
                "ужин",
                "завтрак",
                "перекус",
                "кофе",
                "мартирос",
                "вода",
                "шоколадки",
            ],
            "Транспорт": [
                "такси",
                "метро",
                "автобус",
                "транспорт",
                "дорога",
                "проезд",
                "мытищи",
                "подлипки",
                "маршрутка",
                "электричка",
                "до мытищ",
            ],
            "Жилье": [
                "аренда",
                "квартира",
                "коммуналка",
                "интернет",
                "счета",
                "жкх",
                "ремонт",
            ],
            "Развлечения": [
                "кино",
                "ресторан",
                "кафе",
                "бар",
                "концерт",
                "развлечения",
                "кинотеатр",
                "театр",
                "музей",
            ],
            "Здоровье": [
                "лекарства",
                "врач",
                "аптека",
                "медицина",
                "больница",
                "анализы",
                "стоматолог",
                "окулист",
            ],
            "Одежда": [
                "одежда",
                "обувь",
                "магазин",
                "куртка",
                "брюки",
                "рубашка",
                "платье",
                "кроссовки",
            ],
            "Образование": [
                "курсы",
                "обучение",
                "книги",
                "образование",
                "тренинг",
                "семинар",
                "мастер-класс",
            ],
            "Техника": [
                "техника",
                "гаджеты",
                "электроника",
                "телефон",
                "компьютер",
                "ноутбук",
                "планшет",
            ],
            "Подарки": [
                "подарок",
                "подарки",
                "сувенир",
                "поздравление",
                "праздник",
                "8 марта",
                "пакеты для",
                "шоколадки",
            ],
            "Связь": [
                "телефон",
                "связь",
                "симка",
                "симку",
                "впн",
                "vpn",
                "интернет",
                "роутер",
                "модем",
            ],
        }

        # Состояния для ConversationHandler
        (
            self.CHOOSING_TYPE,
            self.ENTERING_AMOUNT,
            self.ENTERING_DESCRIPTION,
            self.CHOOSING_CATEGORY,
        ) = range(4)

        # Данные для хранения состояния разговора
        self.user_data = {}

    async def middleware_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE, handler
    ):
        """Обработчик для применения всех middleware"""

        # Применяем logging middleware
        async def logging_wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE):
            return await self.logging_middleware(update, context, handler)

        # Применяем metrics middleware
        return await self.metrics_middleware(update, context, logging_wrapped)

    def register_handlers(self, application: Application):
        """Регистрация всех обработчиков команд с применением middleware"""

        def wrap_handler(handler):
            async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE):
                # Применяем logging middleware
                async def handler_with_logging(
                    update: Update, context: ContextTypes.DEFAULT_TYPE
                ):
                    return await handler(update, context)

                # Применяем metrics middleware
                return await self.metrics_middleware(
                    update,
                    context,
                    lambda u, c: self.logging_middleware(u, c, handler_with_logging),
                )

            return wrapped

        # Регистрируем обработчики команд с middleware
        application.add_handler(CommandHandler("start", wrap_handler(self.start)))
        application.add_handler(CommandHandler("help", wrap_handler(self.help)))
        application.add_handler(CommandHandler("balance", wrap_handler(self.balance)))
        application.add_handler(CommandHandler("history", wrap_handler(self.history)))
        application.add_handler(CommandHandler("stats", wrap_handler(self.stats)))
        application.add_handler(CommandHandler("total", wrap_handler(self.total)))
        application.add_handler(CommandHandler("category", wrap_handler(self.category)))
        application.add_handler(CommandHandler("export", wrap_handler(self.export)))
        application.add_handler(CommandHandler("clean_db", wrap_handler(self.clean_db)))

        # Регистрируем обработчик текстовых сообщений
        application.add_handler(
            MessageHandler(
                filters.TEXT & ~filters.COMMAND,
                wrap_handler(self.process_transaction_message),
            )
        )

        # Добавляем обработчик для кнопок
        application.add_handler(CallbackQueryHandler(self.button_handler))

        # Добавляем обработчик для интерактивного добавления транзакций
        conv_handler = ConversationHandler(
            entry_points=[CommandHandler("add", self.add_transaction_start)],
            states={
                self.CHOOSING_TYPE: [CallbackQueryHandler(self.type_choice)],
                self.ENTERING_AMOUNT: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.amount_entered)
                ],
                self.ENTERING_DESCRIPTION: [
                    MessageHandler(
                        filters.TEXT & ~filters.COMMAND, self.description_entered
                    )
                ],
                self.CHOOSING_CATEGORY: [CallbackQueryHandler(self.category_choice)],
            },
            fallbacks=[CommandHandler("cancel", self.cancel_transaction)],
        )
        application.add_handler(conv_handler)

        # Регистрируем обработчик ошибок
        application.add_error_handler(self.error_handler)

    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /start"""
        try:
            # Получаем информацию о пользователе
            user = update.effective_user
            db = SessionLocal()

            # Создаем запись о пользователе, если его нет в базе
            db_user = db.query(User).filter(User.telegram_id == user.id).first()
            if not db_user:
                db_user = User(telegram_id=user.id)
                db.add(db_user)
                db.commit()
                logger.info(LOG_NEW_USER.format(user_id=user.id))

            await update.message.reply_text(START_MESSAGE)

        except Exception as e:
            logger.error(LOG_START_ERROR, exc_info=e)
            await update.message.reply_text(ERROR_GENERAL)
        finally:
            db.close()

    async def help(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /help"""
        await update.message.reply_text(HELP_MESSAGE)

    def parse_amount(self, amount_str: str) -> float:
        """Парсинг суммы с учетом пробелов и других символов"""
        # Удаляем все нецифровые символы, кроме точки, запятой и пробелов
        clean_str = "".join(c for c in amount_str if c.isdigit() or c in "., ")
        # Удаляем все пробелы
        clean_str = clean_str.replace(" ", "")
        # Заменяем запятую на точку
        clean_str = clean_str.replace(",", ".")
        return float(clean_str)

    def split_message(self, message: str) -> tuple[str, str]:
        """Разделяет сообщение на сумму и описание"""
        # Пропускаем знак +/- в начале
        message = message[1:].strip()

        # Ищем первое слово, которое не является частью суммы
        parts = message.split()
        amount_parts = []
        description_parts = []

        for part in parts:
            # Если часть содержит цифры или является обозначением валюты
            if any(c.isdigit() for c in part) or part.lower() in [
                "руб",
                "руб.",
                "р.",
                "рублей",
            ]:
                amount_parts.append(part)
            else:
                description_parts.append(part)
                break

        # Добавляем оставшиеся части к описанию
        description_parts.extend(parts[len(amount_parts) + len(description_parts) :])

        return " ".join(amount_parts), " ".join(description_parts)

    def determine_category(self, description: str) -> str:
        """Определение категории по описанию"""
        # Сначала проверяем специальные случаи
        if "8 марта" in description.lower():
            return "Подарки"

        # Затем проверяем обычные ключевые слова
        for category, keywords in self.category_keywords.items():
            if any(keyword.lower() in description.lower() for keyword in keywords):
                return category
        return CATEGORY_DEFAULT

    async def process_transaction_message(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ):
        """Обработка сообщений о транзакциях в свободной форме"""
        message = update.message.text.strip()
        user_id = update.effective_user.id

        # Проверяем, является ли сообщение расходом или доходом
        is_expense = message.startswith("-")
        is_income = message.startswith("+")

        if not (is_expense or is_income):
            # Не транзакция, игнорируем
            return

        try:
            # Если это расход или доход, парсим сообщение
            amount_str, description = self.split_message(message)

            if not description:
                # Если нет описания, просим добавить его
                await update.message.reply_text(ERROR_NO_DESCRIPTION)
                return

            # Парсим сумму
            try:
                amount = self.parse_amount(amount_str)
            except ValueError:
                await update.message.reply_text(ERROR_INVALID_AMOUNT)
                return

            # Определяем категорию
            category_name = self.determine_category(description)

            # Сохраняем транзакцию в базу данных
            db = SessionLocal()

            # Получаем или создаем пользователя
            user = db.query(User).filter(User.telegram_id == user_id).first()
            if not user:
                user = User(telegram_id=user_id)
                db.add(user)
                db.commit()
                db.refresh(user)

            # Получаем или создаем категорию
            category = db.query(Category).filter(Category.name == category_name).first()
            if not category:
                # Если категория не найдена и это не "Без категории", создаем новую
                if category_name != CATEGORY_DEFAULT:
                    category = Category(name=category_name)
                    db.add(category)
                    db.commit()
                    db.refresh(category)
                else:
                    # Если это "Без категории", получаем существующую или создаем
                    category = (
                        db.query(Category)
                        .filter(Category.name == CATEGORY_DEFAULT)
                        .first()
                    )
                    if not category:
                        category = Category(name=CATEGORY_DEFAULT)
                        db.add(category)
                        db.commit()
                        db.refresh(category)

                # Создаем транзакцию
            transaction_type = (
                TransactionType.EXPENSE if is_expense else TransactionType.INCOME
            )

            transaction = Transaction(
                user_id=user.id,
                amount=amount,
                description=description,
                category_id=category.id,
                type=transaction_type,
                created_at=datetime.now(),
            )

            db.add(transaction)
            db.commit()

            # Предлагаем изменить категорию, если это нужно
            categories = db.query(Category).all()

            # Создаем клавиатуру с кнопками категорий
            keyboard = []
            for i in range(0, len(categories), 2):
                row = []
                for j in range(2):
                    if i + j < len(categories):
                        cat = categories[i + j]
                        # В callback data сохраняем id категории и id транзакции
                        row.append(
                            InlineKeyboardButton(
                                cat.name,
                                callback_data=f"category:{cat.id}:{transaction.id}",
                            )
                        )
                if row:
                    keyboard.append(row)

            db.close()

            reply_markup = InlineKeyboardMarkup(keyboard)

            # Отправляем сообщение с подтверждением
            sign = "-" if transaction_type == TransactionType.EXPENSE else "+"
            await update.message.reply_text(
                ADD_TRANSACTION_SAVED.format(
                    sign=sign,
                    amount=amount,
                    category=category_name,
                    description=description,
                ),
                reply_markup=reply_markup,
            )

        except Exception as e:
            bot_logger.error(f"Ошибка при обработке сообщения о транзакции: {e}")
            await update.message.reply_text(ERROR_PROCESSING)

    async def button_handler(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик нажатий на кнопки"""
        query = update.callback_query
        await query.answer()

        data = query.data.split(":")
        action = data[0]

        if action == "category":
            category_id = int(data[1])
            user_id = int(data[2])
            return await self.set_category_for_transaction(query, category_id, user_id)
        elif action == "clean_db_confirm":
            days = int(data[1])
            # Получаем дату, старше которой будем удалять транзакции
            cutoff_date = datetime.now() - timedelta(days=days)

            db = SessionLocal()
            try:
                # Получаем количество старых транзакций
                old_transactions = (
                    db.query(Transaction)
                    .filter(Transaction.created_at < cutoff_date)
                    .all()
                )

                if not old_transactions:
                    await query.edit_message_text(
                        CLEAN_DB_NO_OLD_TRANSACTIONS.format(days=days)
                    )
                    return

                # Удаляем старые транзакции
                count = len(old_transactions)
                for transaction in old_transactions:
                    db.delete(transaction)
                db.commit()

                await query.edit_message_text(
                    CLEAN_DB_SUCCESS.format(count=count, days=days)
                )
            except Exception as e:
                logger.error(f"Ошибка при очистке БД: {e}")
                await query.edit_message_text(ERROR_GENERAL)
            finally:
                db.close()
        elif action == "clean_db_cancel":
            await query.edit_message_text(CLEAN_DB_CANCELLED)

    async def add_transaction_start(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ):
        """Начало диалога добавления транзакции"""
        keyboard = [
            [
                InlineKeyboardButton("Расход", callback_data="type:expense"),
                InlineKeyboardButton("Доход", callback_data="type:income"),
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            ADD_TRANSACTION_START, reply_markup=reply_markup
        )
        return self.CHOOSING_TYPE

    async def type_choice(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработка выбора типа транзакции"""
        query = update.callback_query
        await query.answer()

        user_id = query.from_user.id
        if user_id not in self.user_data:
            self.user_data[user_id] = {}

        data = query.data.split(":")
        trans_type = data[1]

        self.user_data[user_id]["type"] = trans_type

        await query.edit_message_text(
            ADD_TRANSACTION_AMOUNT.format(
                type="расхода" if trans_type == "expense" else "дохода"
            )
        )
        return self.ENTERING_AMOUNT

    async def amount_entered(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработка ввода суммы"""
        user_id = update.effective_user.id
        text = update.message.text

        try:
            # Очищаем ввод от возможных пробелов и запятых
            amount_text = text.replace(" ", "").replace(",", ".")
            amount = float(amount_text)

            self.user_data[user_id]["amount"] = amount
            await update.message.reply_text(ADD_TRANSACTION_DESCRIPTION)
            return self.ENTERING_DESCRIPTION
        except ValueError:
            await update.message.reply_text(ADD_TRANSACTION_INVALID_AMOUNT)
            return self.ENTERING_AMOUNT

    async def description_entered(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ):
        """Обработка ввода описания"""
        user_id = update.effective_user.id
        description = update.message.text

        self.user_data[user_id]["description"] = description

        # Автоматическое определение категории
        category = CATEGORY_DEFAULT
        for cat, keywords in self.category_keywords.items():
            if any(keyword.lower() in description.lower() for keyword in keywords):
                category = cat
                break

        # Предлагаем пользователю выбрать или подтвердить категорию
        keyboard = []
        row = []

        # Получаем все категории из базы данных
        db = SessionLocal()
        categories = db.query(Category).all()
        db.close()

        for i, cat in enumerate(categories):
            # Создаем кнопки по 3 в ряд
            row.append(
                InlineKeyboardButton(
                    f"{'✓ ' if cat.name == category else ''}{cat.name}",
                    callback_data=f"cat:{cat.id}",
                )
            )

            if (i + 1) % 3 == 0 or i == len(categories) - 1:
                keyboard.append(row)
                row = []

        reply_markup = InlineKeyboardMarkup(keyboard)

        await update.message.reply_text(
            ADD_TRANSACTION_CATEGORY.format(category=category),
            reply_markup=reply_markup,
        )

        return self.CHOOSING_CATEGORY

    async def category_choice(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработка выбора категории"""
        query = update.callback_query
        await query.answer()

        user_id = query.from_user.id
        data = query.data.split(":")
        category_id = int(data[1])

        # Получаем категорию из базы данных
        db = SessionLocal()
        category = db.query(Category).filter(Category.id == category_id).first()

        if not category:
            await query.edit_message_text(ERROR_CATEGORY_NOT_FOUND)
            db.close()
            return self.CHOOSING_CATEGORY

        # Получаем пользователя или создаем нового
        user = db.query(User).filter(User.telegram_id == user_id).first()
        if not user:
            user = User(telegram_id=user_id)
            db.add(user)
            db.commit()
            db.refresh(user)

        # Создаем транзакцию
        transaction_type = (
            TransactionType.EXPENSE
            if self.user_data[user_id]["type"] == "expense"
            else TransactionType.INCOME
        )
        amount = self.user_data[user_id]["amount"]
        description = self.user_data[user_id]["description"]

        transaction = Transaction(
            user_id=user.id,
            amount=amount,
            description=description,
            category_id=category.id,
            type=transaction_type,
            created_at=datetime.now(),
        )

        db.add(transaction)
        db.commit()
        db.close()

        # Очищаем данные пользователя
        del self.user_data[user_id]

        # Отправляем подтверждение
        sign = "-" if transaction_type == TransactionType.EXPENSE else "+"
        await query.edit_message_text(
            ADD_TRANSACTION_SAVED.format(
                sign=sign,
                amount=amount,
                category=category.name,
                description=description,
            )
        )

        return ConversationHandler.END

    async def cancel_transaction(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ):
        """Отмена добавления транзакции"""
        user_id = update.effective_user.id

        if user_id in self.user_data:
            del self.user_data[user_id]

        await update.message.reply_text(ADD_TRANSACTION_CANCELLED)
        return ConversationHandler.END

    async def balance(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Показать текущий баланс пользователя"""
        try:
            db = SessionLocal()
            user = (
                db.query(User)
                .filter(User.telegram_id == update.effective_user.id)
                .first()
            )

            if not user:
                await update.message.reply_text(ERROR_NOT_STARTED)
                return

            # Получаем сумму доходов
            income = (
                db.query(func.sum(Transaction.amount))
                .filter(
                    Transaction.user_id == user.id,
                    Transaction.type == TransactionType.INCOME,
                )
                .scalar()
                or 0
            )

            # Получаем сумму расходов
            expenses = (
                db.query(func.sum(Transaction.amount))
                .filter(
                    Transaction.user_id == user.id,
                    Transaction.type == TransactionType.EXPENSE,
                )
                .scalar()
                or 0
            )

            balance = income - expenses

            message = BALANCE_MESSAGE.format(
                income=income, expenses=expenses, balance=balance
            )

            await update.message.reply_text(message)

        except Exception as e:
            logger.error(LOG_BALANCE_ERROR, exc_info=e)
            await update.message.reply_text(ERROR_GENERAL)
        finally:
            db.close()

    async def history(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Показать историю транзакций"""
        try:
            db = SessionLocal()
            user = (
                db.query(User)
                .filter(User.telegram_id == update.effective_user.id)
                .first()
            )

            if not user:
                await update.message.reply_text(ERROR_NOT_STARTED)
                return

            # Определяем период из аргументов команды
            period = "месяц"  # По умолчанию
            page = 1
            page_size = 10

            if context.args:
                for arg in context.args:
                    if arg.lower() in ["день", "неделя", "месяц", "год"]:
                        period = arg.lower()
                    elif arg.startswith("page="):
                        try:
                            page = int(arg.split("=")[1])
                        except ValueError:
                            pass

            # Определяем начальную дату периода
            now = datetime.utcnow()
            if period == "день":
                start_date = now - timedelta(days=1)
                period_name = PERIOD_DAY
            elif period == "неделя":
                start_date = now - timedelta(weeks=1)
                period_name = PERIOD_WEEK
            elif period == "месяц":
                start_date = now - timedelta(days=30)
                period_name = PERIOD_MONTH
            elif period == "год":
                start_date = now - timedelta(days=365)
                period_name = PERIOD_YEAR

            # Получаем транзакции за период
            transactions = (
                db.query(Transaction)
                .filter(
                    Transaction.user_id == user.id, Transaction.created_at >= start_date
                )
                .order_by(Transaction.created_at.desc())
                .all()
            )

            if not transactions:
                await update.message.reply_text(
                    HISTORY_EMPTY.format(period=period_name)
                )
                return

            # Группируем транзакции по дням
            transactions_by_day = {}
            for t in transactions:
                day = t.created_at.date()
                if day not in transactions_by_day:
                    transactions_by_day[day] = {
                        "transactions": [],
                        "income": 0,
                        "expenses": 0,
                    }
                transactions_by_day[day]["transactions"].append(t)
                if t.type == TransactionType.INCOME:
                    transactions_by_day[day]["income"] += t.amount
                else:
                    transactions_by_day[day]["expenses"] += t.amount

            # Сортируем дни по убыванию
            sorted_days = sorted(transactions_by_day.keys(), reverse=True)

            # Подсчитываем общие суммы за период
            total_income = sum(
                day_data["income"] for day_data in transactions_by_day.values()
            )
            total_expenses = sum(
                day_data["expenses"] for day_data in transactions_by_day.values()
            )

            # Готовим сообщение
            message = HISTORY_HEADER.format(period=period_name)
            message += HISTORY_SUMMARY.format(
                total_income=total_income,
                total_expenses=total_expenses,
                balance=total_income - total_expenses,
            )

            # Пагинация
            total_pages = (len(sorted_days) + page_size - 1) // page_size
            page = min(max(1, page), total_pages)
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            current_days = sorted_days[start_idx:end_idx]

            # Добавляем транзакции по дням
            for day in current_days:
                day_data = transactions_by_day[day]
                message += HISTORY_DAY_HEADER.format(
                    date=day.strftime("%d.%m.%Y"),
                    income=day_data["income"],
                    expenses=day_data["expenses"],
                )

                # Группировка одинаковых транзакций
                grouped_transactions = {}
                for t in day_data["transactions"]:
                    category = t.category.name if t.category else CATEGORY_DEFAULT
                    # Ключ для группировки: тип, сумма, описание, категория
                    key = (t.type, t.amount, t.description, category)
                    if key in grouped_transactions:
                        grouped_transactions[key]["count"] += 1
                    else:
                        grouped_transactions[key] = {
                            "transaction": t,
                            "count": 1,
                            "category": category,
                        }

                # Выводим транзакции (с учетом группировки)
                for key, group in grouped_transactions.items():
                    t = group["transaction"]
                    category = group["category"]
                    count = group["count"]

                    # Если это единичная транзакция
                    if count == 1:
                        message += HISTORY_TRANSACTION.format(
                            emoji="-" if t.type == TransactionType.EXPENSE else "+",
                            amount=t.amount,
                            description=t.description,
                            category=category,
                        )
                    else:
                        # Если это группа одинаковых транзакций
                        message += HISTORY_TRANSACTION_GROUP.format(
                            emoji="—" if t.type == TransactionType.EXPENSE else " +",
                            amount=t.amount,
                            description=t.description,
                            category=category,
                            count=count,
                        )

                # Убираем пустую строку после каждого дня
                if day != current_days[-1]:  # Если не последний день
                    message += "\n"

            # Добавляем информацию о пагинации
            if total_pages > 1:
                message += HISTORY_PAGINATION.format(
                    current_page=page, total_pages=total_pages
                )

            await update.message.reply_text(message)

        except Exception as e:
            logger.error(LOG_HISTORY_ERROR, exc_info=e)
            await update.message.reply_text(ERROR_GENERAL)
        finally:
            db.close()

    async def stats(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Показать статистику за период"""
        try:
            period = "месяц"  # По умолчанию
            if context.args:
                period = context.args[0].lower()

            db = SessionLocal()
            user = (
                db.query(User)
                .filter(User.telegram_id == update.effective_user.id)
                .first()
            )

            if not user:
                await update.message.reply_text(ERROR_NOT_STARTED)
                return

            # Определяем начальную дату периода
            now = datetime.utcnow()
            if period == "день":
                start_date = now - timedelta(days=1)
                period_name = "день"
            elif period == "неделя":
                start_date = now - timedelta(weeks=1)
                period_name = "неделю"
            elif period == "месяц":
                start_date = now - timedelta(days=30)
                period_name = "месяц"
            elif period == "год":
                start_date = now - timedelta(days=365)
                period_name = "год"
            else:
                await update.message.reply_text(SYSTEM_INVALID_PERIOD)
                return

            # Получаем статистику
            income = (
                db.query(func.sum(Transaction.amount))
                .filter(
                    Transaction.user_id == user.id,
                    Transaction.type == TransactionType.INCOME,
                    Transaction.created_at >= start_date,
                )
                .scalar()
                or 0
            )

            expenses = (
                db.query(func.sum(Transaction.amount))
                .filter(
                    Transaction.user_id == user.id,
                    Transaction.type == TransactionType.EXPENSE,
                    Transaction.created_at >= start_date,
                )
                .scalar()
                or 0
            )

            # Топ категорий расходов
            top_expenses = (
                db.query(Category.name, func.sum(Transaction.amount).label("total"))
                .join(Transaction.category)
                .filter(
                    Transaction.user_id == user.id,
                    Transaction.type == TransactionType.EXPENSE,
                    Transaction.created_at >= start_date,
                )
                .group_by(Category.name)
                .order_by(func.sum(Transaction.amount).desc())
                .limit(5)
                .all()
            )

            # Формируем сообщение о категориях
            categories_message = ""
            if top_expenses:
                categories_message = TOP_CATEGORIES_HEADER
                for category, amount in top_expenses:
                    categories_message += TOP_CATEGORY_ITEM.format(
                        category=category, amount=amount
                    )

            message = STATS_MESSAGE.format(
                period=period_name,
                income=income,
                expenses=expenses,
                balance=income - expenses,
                categories=categories_message,
            )

            await update.message.reply_text(message)

        except Exception as e:
            logger.error(LOG_STATS_ERROR, exc_info=e)
            await update.message.reply_text(ERROR_GENERAL)
        finally:
            db.close()

    async def total(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Показать общую сумму расходов и доходов группы"""
        try:
            db = SessionLocal()

            # Получаем общую сумму доходов всех пользователей
            total_income = (
                db.query(func.sum(Transaction.amount))
                .filter(Transaction.type == TransactionType.INCOME)
                .scalar()
                or 0
            )

            # Получаем общую сумму расходов всех пользователей
            total_expenses = (
                db.query(func.sum(Transaction.amount))
                .filter(Transaction.type == TransactionType.EXPENSE)
                .scalar()
                or 0
            )

            # Получаем количество пользователей
            users_count = db.query(User).count()

            # Формируем сообщение
            message = TOTAL_STATS.format(
                users_count=users_count,
                total_income=total_income,
                total_expenses=total_expenses,
                group_balance=total_income - total_expenses,
                avg_income=total_income / users_count if users_count > 0 else 0,
                avg_expenses=total_expenses / users_count if users_count > 0 else 0,
            )

            await update.message.reply_text(message)

        except Exception as e:
            logger.error(LOG_TOTAL_ERROR, exc_info=e)
            await update.message.reply_text(ERROR_GENERAL)
        finally:
            db.close()

    async def category(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Показать статистику по конкретной категории"""
        try:
            if not context.args:
                # Показываем список всех категорий
                db = SessionLocal()
                categories = db.query(Category).all()

                if not categories:
                    await update.message.reply_text(CATEGORY_NO_CATEGORIES)
                    return

                message = CATEGORY_LIST_HEADER
                for cat in categories:
                    message += CATEGORY_LIST_ITEM.format(name=cat.name)

                message += CATEGORY_LIST_FOOTER

                await update.message.reply_text(message)
                return

            # Получаем название категории из аргументов
            category_name = " ".join(context.args)

            db = SessionLocal()
            # Ищем категорию
            category = (
                db.query(Category).filter(Category.name.ilike(category_name)).first()
            )

            if not category:
                await query.edit_message_text(
                    CATEGORY_NOT_FOUND.format(name=category_name)
                )
                db.close()
                return

            # Получаем статистику по категории для текущего пользователя
            user = (
                db.query(User)
                .filter(User.telegram_id == update.effective_user.id)
                .first()
            )

            if not user:
                await update.message.reply_text(
                    "Пожалуйста, запустите бота командой /start"
                )
                return

            # Получаем транзакции пользователя в этой категории
            transactions = (
                db.query(Transaction)
                .filter(
                    Transaction.user_id == user.id,
                    Transaction.category_id == category.id,
                )
                .order_by(Transaction.created_at.desc())
                .all()
            )

            if not transactions:
                await update.message.reply_text(
                    CATEGORY_NO_TRANSACTIONS.format(name=category.name)
                )
                return

            # Считаем статистику
            total_amount = sum(t.amount for t in transactions)
            avg_amount = total_amount / len(transactions)

            # Разделяем на доходы и расходы
            expenses = [t for t in transactions if t.type == TransactionType.EXPENSE]
            incomes = [t for t in transactions if t.type == TransactionType.INCOME]

            total_expenses = sum(t.amount for t in expenses)
            total_incomes = sum(t.amount for t in incomes)

            message = CATEGORY_STATS.format(
                name=category.name,
                total=len(transactions),
                total_amount=total_amount,
                avg_amount=avg_amount,
                expenses=total_expenses,
                expense_count=len(expenses),
                incomes=total_incomes,
                income_count=len(incomes),
            )

            # Добавляем последние 5 операций
            for t in transactions[:5]:
                operation = "📉" if t.type == TransactionType.EXPENSE else "📈"
                message += CATEGORY_TRANSACTION.format(
                    emoji=operation,
                    date=t.created_at.strftime("%d.%m.%Y"),
                    amount=t.amount,
                    description=t.description,
                )

            await update.message.reply_text(message)

        except Exception as e:
            logger.error(LOG_CATEGORY_ERROR, exc_info=e)
            await update.message.reply_text(
                "Произошла ошибка при получении статистики по категории"
            )
        finally:
            db.close()

    async def export(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Экспорт истории транзакций в Excel"""
        try:
            db = SessionLocal()
            user = (
                db.query(User)
                .filter(User.telegram_id == update.effective_user.id)
                .first()
            )

            if not user:
                await update.message.reply_text(
                    "Пожалуйста, запустите бота командой /start"
                )
                return

            # Получаем все транзакции пользователя
            transactions = (
                db.query(Transaction)
                .filter(Transaction.user_id == user.id)
                .order_by(Transaction.created_at.desc())
                .all()
            )

            if not transactions:
                await update.message.reply_text(EXPORT_EMPTY)
                return

            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Транзакции"

            # Определяем стили
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Записываем заголовки
            for col, header in enumerate(EXPORT_HEADERS, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            # Записываем данные
            for row, t in enumerate(transactions, 2):
                # Дата
                ws.cell(
                    row=row, column=1, value=t.created_at.strftime("%d.%m.%Y %H:%M")
                )

                # Тип
                transaction_type = (
                    EXPORT_TRANSACTION_TYPE_EXPENSE
                    if t.type == TransactionType.EXPENSE
                    else EXPORT_TRANSACTION_TYPE_INCOME
                )
                ws.cell(row=row, column=2, value=transaction_type)

                # Сумма с форматированием в зависимости от типа
                amount_cell = ws.cell(row=row, column=3, value=t.amount)
                if t.type == TransactionType.EXPENSE:
                    amount_cell.font = Font(color="FF0000")  # Красный для расходов
                else:
                    amount_cell.font = Font(color="008000")  # Зеленый для доходов

                # Описание
                ws.cell(row=row, column=4, value=t.description)

                # Категория
                ws.cell(
                    row=row,
                    column=5,
                    value=t.category.name if t.category else CATEGORY_DEFAULT,
                )

                # Добавляем границы для всех ячеек
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border

            # Автоматическая ширина столбцов
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

            # Сохраняем в буфер
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Отправляем файл
            await update.message.reply_document(
                document=excel_buffer,
                filename=f"transactions_{datetime.now().strftime('%Y%m%d')}.xlsx",
                caption=EXPORT_CAPTION,
            )

        except Exception as e:
            logger.error(LOG_TRANSACTION_ERROR, exc_info=e)
            await update.message.reply_text(ERROR_GENERAL)
        finally:
            db.close()

    async def error_handler(
        self, update: Update, context: ContextTypes.DEFAULT_TYPE
    ) -> None:
        """Обработчик ошибок бота"""
        logger.error(
            LOG_UPDATE_ERROR.format(update=update, error=context.error),
            exc_info=context.error,
        )

        try:
            if update and update.effective_message:
                await update.effective_message.reply_text(ERROR_BOT)

        except Exception as e:
            logger.error(LOG_ERROR_MESSAGE_ERROR.format(error=e), exc_info=True)

    async def set_category_for_transaction(self, query, category_id, transaction_id):
        """Устанавливает категорию для транзакции после выбора пользователем"""
        db = None
        try:
            db = SessionLocal()
            transaction = (
                db.query(Transaction).filter(Transaction.id == transaction_id).first()
            )
            category = db.query(Category).filter(Category.id == category_id).first()

            if not transaction or not category:
                await query.edit_message_text(ERROR_CATEGORY_NOT_FOUND)
                return

            # Обновляем категорию
            transaction.category_id = category.id
            db.commit()

            # Получаем обновленную транзакцию для отображения
            transaction = (
                db.query(Transaction).filter(Transaction.id == transaction_id).first()
            )
            sign = "-" if transaction.type == TransactionType.EXPENSE else "+"

            await query.edit_message_text(
                CHANGE_CATEGORY_SUCCESS.format(
                    sign=sign,
                    amount=transaction.amount,
                    category=category.name,
                    description=transaction.description,
                )
            )

        except Exception as e:
            bot_logger.error(LOG_CATEGORY_CHANGE_ERROR.format(error=e))
            await query.edit_message_text(ERROR_GENERAL)
        finally:
            if db:
                db.close()

    async def clean_db(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Очистка старых транзакций из БД (только для администраторов)"""
        user_id = update.effective_user.id

        # Проверяем, является ли пользователь администратором
        if user_id not in self.admin_ids:
            await update.message.reply_text(CLEAN_DB_NOT_ADMIN)
            return

        # По умолчанию удаляем транзакции старше 90 дней
        days = 90
        if context.args and context.args[0].isdigit():
            days = int(context.args[0])

        # Создаем кнопки подтверждения
        keyboard = [
            [
                InlineKeyboardButton(
                    "✅ Подтвердить", callback_data=f"clean_db_confirm:{days}"
                ),
                InlineKeyboardButton("❌ Отмена", callback_data="clean_db_cancel"),
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await update.message.reply_text(
            CLEAN_DB_CONFIRM.format(days=days), reply_markup=reply_markup
        )

    def run(self):
        """Запуск бота"""
        # Создаем приложение
        application = Application.builder().token(self.token).build()

        # Регистрируем обработчики
        self.register_handlers(application)

        # Запускаем бота
        application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    bot = FinanceBot()
    bot.run()
